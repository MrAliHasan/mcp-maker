"""
MCP-Maker Excel Connector — Inspect Excel (.xlsx) files.

Each sheet in the workbook becomes a table, with the first row as column headers.
"""

import datetime
import os

from ..core.schema import (
    Column,
    ColumnType,
    DataSourceSchema,
    Table,
)
from .base import BaseConnector, register_connector
from .utils import sanitize_name


# Map Python/openpyxl types to our universal types
def _infer_column_type(values: list) -> ColumnType:
    """Infer column type from sample values."""
    non_none = [v for v in values if v is not None]
    if not non_none:
        return ColumnType.STRING

    type_counts = {"int": 0, "float": 0, "bool": 0, "str": 0, "datetime": 0, "date": 0}
    for v in non_none[:100]:  # Sample first 100 rows
        if isinstance(v, bool):
            type_counts["bool"] += 1
        elif isinstance(v, datetime.datetime):
            type_counts["datetime"] += 1
        elif isinstance(v, datetime.date):
            type_counts["date"] += 1
        elif isinstance(v, int):
            type_counts["int"] += 1
        elif isinstance(v, float):
            type_counts["float"] += 1
        else:
            type_counts["str"] += 1

    # A mix of ints and floats is a float column, not "whichever won the vote"
    if type_counts["float"] and type_counts["int"] and not (type_counts["str"] or type_counts["bool"]):
        return ColumnType.FLOAT

    dominant = max(type_counts, key=type_counts.get)
    return {
        "int": ColumnType.INTEGER,
        "float": ColumnType.FLOAT,
        "bool": ColumnType.BOOLEAN,
        "str": ColumnType.STRING,
        "datetime": ColumnType.DATETIME,
        "date": ColumnType.DATE,
    }[dominant]


def _dedupe(names: list[str]) -> list[str]:
    """Make sanitized names unique by appending _2, _3, … to duplicates."""
    seen: dict[str, int] = {}
    result = []
    for name in names:
        if name in seen:
            seen[name] += 1
            result.append(f"{name}_{seen[name]}")
        else:
            seen[name] = 1
            result.append(name)
    return result


class ExcelConnector(BaseConnector):
    """Connector for Excel (.xlsx) files.

    Inspects all sheets, treating each sheet as a table.
    The first row is used as column headers.

    URI format: excel:///path/to/file.xlsx or just ./file.xlsx
    """

    @property
    def source_type(self) -> str:
        return "excel"

    def _get_file_path(self) -> str:
        """Extract the file path from the URI."""
        path = self.uri
        if path.startswith("excel:///"):
            path = path[len("excel:///"):]
        elif path.startswith("excel://"):
            path = path[len("excel://"):]
        return os.path.expanduser(path)

    def validate(self) -> bool:
        """Check that the Excel file exists and is readable."""
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel support. "
                "Install it with: pip install mcp-maker[excel]"
            )

        file_path = self._get_file_path()
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        ext = os.path.splitext(file_path)[1].lower()
        if ext not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
            raise ValueError(f"Unsupported file format: {ext}. Use .xlsx files.")

        return True

    def inspect(self) -> DataSourceSchema:
        """Inspect the Excel file and return its schema."""
        import openpyxl

        file_path = self._get_file_path()
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        tables = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Sample up to 500 rows for type inference instead of loading
            # the whole sheet into memory; count the rest without storing.
            rows = []
            total_rows = 0
            for row in ws.iter_rows(values_only=True):
                total_rows += 1
                if total_rows <= 501:  # header + 500 data rows
                    rows.append(row)

            if not rows:
                continue

            # First row = headers
            headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
            data_rows = rows[1:]

            # Skip empty sheets (no headers)
            if not any(h for h in headers):
                continue

            # Infer column types from data (sanitized + de-duplicated headers)
            safe_names = _dedupe([sanitize_name(h) for h in headers])
            columns = []
            for col_idx, (header, safe_name) in enumerate(zip(headers, safe_names)):
                col_values = [row[col_idx] if col_idx < len(row) else None for row in data_rows]
                col_type = _infer_column_type(col_values)
                columns.append(Column(
                    name=safe_name,
                    type=col_type,
                    nullable=True,
                    description=header if header != safe_name else None,
                ))

            safe_sheet = sanitize_name(sheet_name)
            tables.append(Table(
                name=safe_sheet,
                columns=columns,
                row_count=total_rows - 1,
                description=sheet_name,
            ))

        wb.close()

        # De-duplicate sheet names that collide after sanitization
        for table, unique in zip(tables, _dedupe([t.name for t in tables])):
            table.name = unique

        return DataSourceSchema(
            source_type="excel",
            source_uri=self.uri,
            tables=tables,
            metadata={
                "file_path": file_path,
                "file_size_bytes": os.path.getsize(file_path),
                "sheet_count": len(tables),
            },
        )


# Register this connector
register_connector("excel", ExcelConnector)
